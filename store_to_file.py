import os
import locale
import pandas as pd
from PIL import Image

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# def extract_amazon_orders(html_content):
#     soup = BeautifulSoup(html_content, "html.parser")

#     # Hier musst du den HTML-Code analysieren und die entsprechenden Elemente identifizieren
#     # Dies ist ein allgemeines Beispiel und sollte an deine spezielle Situation angepasst werden

#     order_elements = soup.find_all("div", class_="your-order")

#     orders_data = []

#     for order_element in order_elements:
#         # Hier kannst du Informationen aus jedem 'order_element' extrahieren
#         # Beispiel: order_date = order_element.find('span', class_='order-date').text
#         # Füge die extrahierten Daten in die 'orders_data'-Liste ein
#         print("Test")

#     return orders_data

# def save_to_excel(product_list):
#     # Erstelle ein leeres DataFrame
#     df = pd.DataFrame()

#     # Fülle das DataFrame mit den Produktinformationen
#     for product in product_list:
#         df = pd.concat([
#             df,
#             pd.DataFrame({
#                 "Bestelldatum": [product.order_date],
#                 "Lieferdatum": [product.delivery_date],
#                 "Bestellnummer": [product.order_number],
#                 "Produktname": [product.product_name],
#                 "Preis": [product.product_price],
#                 "Bild heruntergeladen": [product.image_filename],
#             })
#         ], ignore_index=True)
#     # Speichere das DataFrame in einer Excel-Datei
#     df.to_excel("amazon_shopping_history.xlsx", index=False)


# def save_to_excel(product_list):
#     # Erstelle ein leeres Excel-Arbeitsblatt
#     wb = Workbook()
#     ws = wb.active
#     current_directory = os.getcwd()


#     # Fülle das Arbeitsblatt mit den Produktinformationen und füge Bilder hinzu
#     for i, product in enumerate(product_list, start=2):
#         ws.cell(row=i, column=1, value=product.order_date)
#         ws.cell(row=i, column=2, value=product.delivery_date)
#         ws.cell(row=i, column=3, value=product.order_number)
#         ws.cell(row=i, column=4, value=product.product_name)
#         ws.cell(row=i, column=5, value=product.product_price)

#         # Füge das Bild zum Arbeitsblatt hinzu
#         img = Image(f"{current_directory}/img/{product.image_filename}")
#         ws.add_image(img, f'F{i}')  # 'F{i}' ist der Zellbezug, wo das Bild eingefügt werden soll

#     # Speichere das Arbeitsblatt in einer Excel-Datei
#     wb.save("amazon_shopping_history.xlsx")


# def image_to_bytes(image_path):
#     with open(image_path, "rb") as f:
#         return BytesIO(f.read())


# def save_to_excel(product_list, file, lang="de"):
#     # Überprüfen Sie, ob die Datei bereits existiert
#     if os.path.exists(file):
#         # Laden Sie das vorhandene Excel-File
#         df_existing = pd.read_excel(file, engine="openpyxl")

#         # Überprüfen Sie die Eindeutigkeit der Bestellnummern in der vorhandenen Datei
#         if (
#             df_existing["Bestellnummer"]
#             .isin([product.order_number for product in product_list])
#             .any()
#         ):
#             print("Es gibt bereits Einträge mit diesen Bestellnummern.")
#             return

#         # Fügen Sie die neuen Daten zum vorhandenen DataFrame hinzu
#         df_new = pd.DataFrame(
#             {
#                 "Bestelldatum": [product.order_date for product in product_list],
#                 "Lieferdatum": [product.delivery_date for product in product_list],
#                 "Bestellnummer": [product.order_number for product in product_list],
#                 "Produktname": [product.product_name for product in product_list],
#                 "Preis": [product.product_price for product in product_list],
#                 "Bild": [product.image_filename for product in product_list],
#             }
#         )

#         df_new["Bestelldatum"] = pd.to_datetime(df_new["Bestelldatum"], dayfirst=True)
#         df_new["Preis"] = df_new["Preis"].str.replace("€", "")
#         df_new["Preis"] = df_new["Preis"].str.replace(",", ".")
#         df_new["Preis"] = pd.to_numeric(df_new["Preis"])

#         # Verknüpfen Sie das bestehende und das neue DataFrame nach Bestellnummer
#         df_combined = pd.concat([df_existing, df_new], ignore_index=True)

#         # Überprüfen Sie die Eindeutigkeit der Bestellnummern im kombinierten DataFrame
#         if df_combined["Bestellnummer"].duplicated().any():
#             print("Es gibt Duplikate in den Bestellnummern.")
#             return

#         # Speichern Sie das kombinierte DataFrame zurück in die Excel-Datei
#         df_combined.to_excel(file, index=False, engine="openpyxl")
#         path_to_img(df=df_combined, file=file)

#     else:
#         # Wenn die Datei nicht existiert, erstellen Sie ein neues Excel-File
#         save_to_excel_new(product_list, file)


# def save_to_excel_new(product_list, file, lang):
#     # Erstelle ein leeres DataFrame
#     df = pd.DataFrame(
#         {
#             "Bestelldatum": [product.order_date for product in product_list],
#             "Lieferdatum": [product.delivery_date for product in product_list],
#             "Bestellnummer": [product.order_number for product in product_list],
#             "Produktname": [product.product_name for product in product_list],
#             "Preis": [product.product_price for product in product_list],
#             "Bild": [product.image_filename for product in product_list],
#         }
#     )
#     df["Bestelldatum"] = pd.to_datetime(df["Bestelldatum"], dayfirst=True)
#     # Annahme: df ist Ihr DataFrame und 'Preisspalte' ist die Spalte mit den Geldbeträgen
#     df["Preis"] = df["Preis"].str.replace("€", "")
#     df["Preis"] = df["Preis"].str.replace(",", ".")
#     df["Preis"] = pd.to_numeric(df["Preis"])

#     # Setzen Sie die locale auf Deutsch für die Währungsformatierung
#     locale.setlocale(locale.LC_ALL, lang)
#     df["Preis"] = df["Preis"].apply(locale.currency)

#     # Speichere das DataFrame in einer Excel-Datei
#     df.to_excel(file, index=False, engine="openpyxl")
#     path_to_img(df=df, file=file)


def save_to_excel(product_list, excel_file, lang="de"):
    # Überprüfen, ob die Excel-Datei bereits existiert
    if os.path.exists(excel_file):
        # Lade vorhandene Daten aus der Excel-Datei
        existing_df = pd.read_excel(excel_file)
        # Überprüfe Duplikate basierend auf der Bestellnummer
        unique_order_numbers = set(existing_df["Bestellnummer"])
    else:
        # Erstelle ein leeres DataFrame, wenn die Excel-Datei nicht existiert
        existing_df = pd.DataFrame()
        unique_order_numbers = set()

    # Erstelle ein leeres DataFrame für die neuen Daten
    new_data = {
        "Bestelldatum": [product.order_date for product in product_list],
        "Lieferdatum": [product.delivery_date for product in product_list],
        "Bestellnummer": [product.order_number for product in product_list],
        "Produktname": [product.product_name for product in product_list],
        "Preis": [product.product_price for product in product_list],
        "Bild": [product.image_filename for product in product_list],
    }

    # Überprüfe Duplikate basierend auf der Bestellnummer, füge nur ein, was eindeutig ist
    new_data = {
        col: [
            val
            for idx, val in enumerate(new_data[col])
            if product_list[idx].order_number not in unique_order_numbers
        ]
        for col in new_data
    }

    # Setzen Sie die locale auf Deutsch für die Währungsformatierung
    locale.setlocale(locale.LC_ALL, lang)

    # Füge neue Daten zum bestehenden DataFrame hinzu
    df = pd.DataFrame(new_data)

    # Definiere das Datumsformat
    date_format = "%d. %B %Y"  # "18. Dezember 2022"
    df["Bestelldatum"] = pd.to_datetime(df["Bestelldatum"], format=date_format, dayfirst=True)

    df["Preis"] = df["Preis"].apply(lambda x: str(x).replace("€", "") if isinstance(x, str) else x)
    df["Preis"] = df["Preis"].apply(lambda x: str(x).replace(",", ".") if isinstance(x, str) else x)
    df["Preis"] = pd.to_numeric(df["Preis"])

    # df["Preis"] = df["Preis"].apply(locale.currency)

    updated_df = pd.concat([existing_df, df], ignore_index=True)

    # Speichere das aktualisierte DataFrame in einer Excel-Datei
    try:
        updated_df.to_excel(excel_file, index=False, engine="openpyxl")
        path_to_img(df=updated_df, file=excel_file)
    except PermissionError as e:
        print(str(e))
    finally:
        return


def path_to_img(df, file):
    # Öffnen Sie die vorhandene Excel-Datei
    workbook = load_workbook(file)

    # Greifen Sie auf das gewünschte Arbeitsblatt zu
    sheet = workbook.active

    # Füge die Bilder in die entsprechenden Zellen ein
    for r_idx, image_path in enumerate(df["Bild"], start=2):
        try:
            img = Image(image_path)
            sheet.add_image(img, f"F{r_idx}")

            # Entferne den Zellwert in der Spalte "Bild"
            sheet.cell(row=r_idx, column=df.columns.get_loc("Bild") + 1, value="")
            sheet.row_dimensions[r_idx].height = img.height * 0.85
            sheet.column_dimensions["F"].width = img.width / 7

        except:
            pass

    # Speichere das Excel-Blatt
    workbook.save(file)
    return
